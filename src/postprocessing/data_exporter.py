"""
DataExporter - Exportador de datos a formatos CSV, Excel, JSON.
"""

import os
import json
import csv
from typing import List, Dict, Any
from datetime import datetime
from ..domain.analysis_results import AnalysisResults


class DataExporter:
    """Exporta datos de análisis a diferentes formatos."""
    
    def __init__(self, output_dir: str):
        """
        Inicializa el exportador de datos.
        
        Args:
            output_dir: Directorio de salida
        """
        self.output_dir = output_dir
        self.ensure_output_dir()
    
    def ensure_output_dir(self):
        """Asegura que el directorio de salida existe."""
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Crear subdirectorio para exports
        exports_dir = os.path.join(self.output_dir, "exports")
        os.makedirs(exports_dir, exist_ok=True)
    
    def export_data(self,
                   analysis_results: AnalysisResults,
                   config: Dict[str, Any] = None) -> List[str]:
        """
        Exporta datos de análisis a diferentes formatos.
        
        Args:
            analysis_results: Resultados del análisis
            config: Configuración de exportación
            
        Returns:
            Lista de archivos exportados
        """
        config = config or {}
        exported_files = []
        
        print(f"💾 Exportando datos para {analysis_results.model_name}")
        
        # 1. Exportar a CSV
        if config.get('export_csv', True):
            csv_files = self._export_to_csv(analysis_results, config)
            exported_files.extend(csv_files)
        
        # 2. Exportar a JSON
        if config.get('export_json', True):
            json_file = self._export_to_json(analysis_results, config)
            if json_file:
                exported_files.append(json_file)
        
        # 3. Exportar a Excel (si está disponible)
        if config.get('export_excel', False):
            excel_file = self._export_to_excel(analysis_results, config)
            if excel_file:
                exported_files.append(excel_file)
        
        return exported_files
    
    def _export_to_csv(self, analysis_results: AnalysisResults, config: Dict[str, Any]) -> List[str]:
        """Exporta datos a archivos CSV."""
        exported_files = []
        exports_dir = os.path.join(self.output_dir, "exports")
        
        try:
            # CSV de resumen general
            summary_file = os.path.join(exports_dir, f"summary_{analysis_results.model_name}.csv")
            self._write_summary_csv(analysis_results, summary_file)
            exported_files.append(summary_file)
            print(f"   ✅ CSV resumen: {os.path.basename(summary_file)}")
            
            # CSV de resultados estáticos
            if analysis_results.static_results:
                static_file = os.path.join(exports_dir, f"static_{analysis_results.model_name}.csv")
                self._write_static_csv(analysis_results.static_results, static_file)
                exported_files.append(static_file)
                print(f"   ✅ CSV estático: {os.path.basename(static_file)}")
            
            # CSV de resultados modales
            if analysis_results.modal_results:
                modal_file = os.path.join(exports_dir, f"modal_{analysis_results.model_name}.csv")
                self._write_modal_csv(analysis_results.modal_results, modal_file)
                exported_files.append(modal_file)
                print(f"   ✅ CSV modal: {os.path.basename(modal_file)}")
            
        except Exception as e:
            print(f"   ❌ Error exportando CSV: {e}")
        
        return exported_files
    
    def _export_to_json(self, analysis_results: AnalysisResults, config: Dict[str, Any]) -> str:
        """Exporta datos a JSON estructurado."""
        try:
            exports_dir = os.path.join(self.output_dir, "exports")
            json_file = os.path.join(exports_dir, f"data_{analysis_results.model_name}.json")
            
            export_data = {
                "export_metadata": {
                    "exported_at": datetime.now().isoformat(),
                    "model_name": analysis_results.model_name,
                    "export_version": "1.0.0"
                },
                "analysis_data": self._prepare_json_data(analysis_results)
            }
            
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            print(f"   ✅ JSON export: {os.path.basename(json_file)}")
            return json_file
            
        except Exception as e:
            print(f"   ❌ Error exportando JSON: {e}")
            return None
    
    def _export_to_excel(self, analysis_results: AnalysisResults, config: Dict[str, Any]) -> str:
        """Exporta datos a Excel (requiere openpyxl)."""
        try:
            import openpyxl
            import openpyxl.styles
            
            exports_dir = os.path.join(self.output_dir, "exports")
            excel_file = os.path.join(exports_dir, f"data_{analysis_results.model_name}.xlsx")
            
            wb = openpyxl.Workbook()
            
            # Hoja de resumen
            ws_summary = wb.active
            ws_summary.title = "Resumen"
            self._write_excel_summary(ws_summary, analysis_results)
            
            # Hoja de resultados estáticos
            if analysis_results.static_results:
                ws_static = wb.create_sheet("Estatico")
                self._write_excel_static(ws_static, analysis_results.static_results)
            
            # Hoja de resultados modales
            if analysis_results.modal_results:
                ws_modal = wb.create_sheet("Modal")
                self._write_excel_modal(ws_modal, analysis_results.modal_results)
            
            wb.save(excel_file)
            print(f"   ✅ Excel export: {os.path.basename(excel_file)}")
            return excel_file
            
        except ImportError:
            print("   ⚠️ openpyxl no disponible, saltando exportación a Excel")
            return None
        except Exception as e:
            print(f"   ❌ Error exportando Excel: {e}")
            return None
    
    def _write_summary_csv(self, analysis_results: AnalysisResults, filename: str):
        """Escribe CSV de resumen."""
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Encabezados
            writer.writerow(['Parámetro', 'Valor'])
            
            # Datos básicos
            writer.writerow(['Modelo', analysis_results.model_name])
            writer.writerow(['Timestamp', analysis_results.timestamp])
            writer.writerow(['Éxito', 'Sí' if analysis_results.success else 'No'])
            writer.writerow(['Tiempo Total (s)', f"{analysis_results.total_analysis_time:.3f}"])
            
            # Resultados estáticos
            if analysis_results.static_results:
                writer.writerow(['--- Análisis Estático ---', ''])
                writer.writerow(['Desplazamiento Máximo (m)', f"{analysis_results.static_results.max_displacement:.6f}"])
                writer.writerow(['Esfuerzo Máximo (MPa)', f"{analysis_results.static_results.max_stress:.2f}"])
                writer.writerow(['Convergencia', 'Sí' if analysis_results.static_results.convergence_achieved else 'No'])
                writer.writerow(['Iteraciones', str(analysis_results.static_results.num_iterations)])
                writer.writerow(['Tiempo Estático (s)', f"{analysis_results.static_results.analysis_time:.3f}"])
            
            # Resultados modales
            if analysis_results.modal_results:
                writer.writerow(['--- Análisis Modal ---', ''])
                writer.writerow(['Número de Modos', str(len(analysis_results.modal_results.periods))])
                if analysis_results.modal_results.periods:
                    writer.writerow(['Primer Período (s)', f"{analysis_results.modal_results.periods[0]:.4f}"])
                writer.writerow(['Tiempo Modal (s)', f"{analysis_results.modal_results.analysis_time:.3f}"])
            
            # Errores
            if analysis_results.errors:
                writer.writerow(['--- Errores ---', ''])
                for i, error in enumerate(analysis_results.errors, 1):
                    writer.writerow([f'Error {i}', error])
    
    def _write_static_csv(self, static_results, filename: str):
        """Escribe CSV de resultados estáticos."""
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Encabezados
            writer.writerow([
                'Parámetro',
                'Valor',
                'Unidad'
            ])
            
            # Datos
            writer.writerow(['Desplazamiento Máximo', f"{static_results.max_displacement:.6f}", 'm'])
            writer.writerow(['Esfuerzo Máximo', f"{static_results.max_stress:.2f}", 'MPa'])
            writer.writerow(['Convergencia Lograda', 'Sí' if static_results.convergence_achieved else 'No', '-'])
            writer.writerow(['Número de Iteraciones', str(static_results.num_iterations), '-'])
            writer.writerow(['Tiempo de Análisis', f"{static_results.analysis_time:.3f}", 's'])
            
            # Si hay desplazamientos nodales
            if hasattr(static_results, 'node_displacements') and static_results.node_displacements:
                writer.writerow(['', '', ''])
                writer.writerow(['--- Desplazamientos Nodales ---', '', ''])
                writer.writerow(['Nodo', 'Dx (m)', 'Dy (m)', 'Dz (m)'])
                
                for node_id, displ in static_results.node_displacements.items():
                    writer.writerow([node_id, f"{displ[0]:.6f}", f"{displ[1]:.6f}", f"{displ[2]:.6f}"])
    
    def _write_modal_csv(self, modal_results, filename: str):
        """Escribe CSV de resultados modales."""
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Encabezados
            writer.writerow([
                'Modo',
                'Período (s)',
                'Frecuencia (Hz)'
            ])
            
            # Datos de modos
            for i, (period, freq) in enumerate(zip(modal_results.periods, modal_results.frequencies), 1):
                writer.writerow([i, f"{period:.4f}", f"{freq:.2f}"])
    
    def _prepare_json_data(self, analysis_results: AnalysisResults) -> Dict[str, Any]:
        """Prepara datos para exportación JSON."""
        data = {
            "model_name": analysis_results.model_name,
            "timestamp": analysis_results.timestamp,
            "success": analysis_results.success,
            "total_analysis_time": analysis_results.total_analysis_time,
            "errors": analysis_results.errors
        }
        
        if analysis_results.static_results:
            data["static_results"] = {
                "max_displacement": analysis_results.static_results.max_displacement,
                "max_stress": analysis_results.static_results.max_stress,
                "convergence_achieved": analysis_results.static_results.convergence_achieved,
                "num_iterations": analysis_results.static_results.num_iterations,
                "analysis_time": analysis_results.static_results.analysis_time
            }
            
            # Agregar desplazamientos nodales si están disponibles
            if hasattr(analysis_results.static_results, 'node_displacements'):
                data["static_results"]["node_displacements"] = analysis_results.static_results.node_displacements
        
        if analysis_results.modal_results:
            data["modal_results"] = {
                "num_modes": len(analysis_results.modal_results.periods),
                "periods": analysis_results.modal_results.periods,
                "frequencies": analysis_results.modal_results.frequencies,
                "analysis_time": analysis_results.modal_results.analysis_time
            }
        
        return data
    
    def _write_excel_summary(self, worksheet, analysis_results: AnalysisResults):
        """Escribe hoja de resumen en Excel."""
        try:
            import openpyxl.styles
            
            # Título
            worksheet['A1'] = 'Resumen de Análisis'
            worksheet['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            
            # Datos básicos
            row = 3
            data_pairs = [
                ('Modelo', analysis_results.model_name),
                ('Timestamp', analysis_results.timestamp),
                ('Éxito', 'Sí' if analysis_results.success else 'No'),
                ('Tiempo Total (s)', f"{analysis_results.total_analysis_time:.3f}")
            ]
            
            for label, value in data_pairs:
                worksheet[f'A{row}'] = label
                worksheet[f'B{row}'] = value
                row += 1
            
            # Resultados estáticos
            if analysis_results.static_results:
                row += 1
                worksheet[f'A{row}'] = 'Análisis Estático'
                worksheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                row += 1
                
                static_data = [
                    ('Desplazamiento Máximo (m)', f"{analysis_results.static_results.max_displacement:.6f}"),
                    ('Convergencia', 'Sí' if analysis_results.static_results.convergence_achieved else 'No')
                ]
                
                for label, value in static_data:
                    worksheet[f'A{row}'] = label
                    worksheet[f'B{row}'] = value
                    row += 1
            
            # Resultados modales
            if analysis_results.modal_results:
                row += 1
                worksheet[f'A{row}'] = 'Análisis Modal'
                worksheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                row += 1
                
                modal_data = [
                    ('Número de Modos', str(len(analysis_results.modal_results.periods))),
                    ('Primer Período (s)', f"{analysis_results.modal_results.periods[0]:.4f}" if analysis_results.modal_results.periods else "N/A")
                ]
                
                for label, value in modal_data:
                    worksheet[f'A{row}'] = label
                    worksheet[f'B{row}'] = value
                    row += 1
        except Exception as e:
            print(f"   ⚠️ Error en formato Excel: {e}")
            # Escribir datos básicos sin formato
            worksheet['A1'] = 'Resumen de Análisis'
            worksheet['A3'] = 'Modelo'
            worksheet['B3'] = analysis_results.model_name
    
    def _write_excel_static(self, worksheet, static_results):
        """Escribe hoja de resultados estáticos en Excel."""
        try:
            import openpyxl.styles
            
            worksheet['A1'] = 'Resultados de Análisis Estático'
            worksheet['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            
            # Encabezados
            headers = ['Parámetro', 'Valor', 'Unidad']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Datos
            data_rows = [
                ('Desplazamiento Máximo', f"{static_results.max_displacement:.6f}", 'm'),
                ('Esfuerzo Máximo', f"{static_results.max_stress:.2f}", 'MPa'),
                ('Convergencia', 'Sí' if static_results.convergence_achieved else 'No', '-'),
                ('Iteraciones', str(static_results.num_iterations), '-'),
                ('Tiempo', f"{static_results.analysis_time:.3f}", 's')
            ]
            
            for row_idx, (param, value, unit) in enumerate(data_rows, 4):
                worksheet.cell(row=row_idx, column=1, value=param)
                worksheet.cell(row=row_idx, column=2, value=value)
                worksheet.cell(row=row_idx, column=3, value=unit)
        except Exception as e:
            print(f"   ⚠️ Error en formato Excel estático: {e}")
            # Escribir datos básicos sin formato
            worksheet['A1'] = 'Resultados de Análisis Estático'
            worksheet['A3'] = 'Desplazamiento Máximo'
            worksheet['B3'] = f"{static_results.max_displacement:.6f}"
    
    def _write_excel_modal(self, worksheet, modal_results):
        """Escribe hoja de resultados modales en Excel."""
        try:
            import openpyxl.styles
            
            worksheet['A1'] = 'Resultados de Análisis Modal'
            worksheet['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            
            # Encabezados
            headers = ['Modo', 'Período (s)', 'Frecuencia (Hz)']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Datos de modos
            for row_idx, (period, freq) in enumerate(zip(modal_results.periods, modal_results.frequencies), 4):
                worksheet.cell(row=row_idx, column=1, value=row_idx - 3)
                worksheet.cell(row=row_idx, column=2, value=f"{period:.4f}")
                worksheet.cell(row=row_idx, column=3, value=f"{freq:.2f}")
        except Exception as e:
            print(f"   ⚠️ Error en formato Excel modal: {e}")
            # Escribir datos básicos sin formato
            worksheet['A1'] = 'Resultados de Análisis Modal'
            if modal_results.periods:
                worksheet['A3'] = 'Primer Período'
                worksheet['B3'] = f"{modal_results.periods[0]:.4f}"
    
    def export_consolidated_data(self, 
                                results_list: List[AnalysisResults],
                                config: Dict[str, Any] = None) -> List[str]:
        """
        Exporta datos consolidados de múltiples análisis.
        
        Args:
            results_list: Lista de resultados de análisis
            config: Configuración de exportación
            
        Returns:
            Lista de archivos exportados
        """
        config = config or {}
        exported_files = []
        
        print(f"📊 Exportando datos consolidados para {len(results_list)} modelos")
        
        try:
            exports_dir = os.path.join(self.output_dir, "exports")
            
            # CSV consolidado
            consolidated_csv = os.path.join(exports_dir, "consolidated_data.csv")
            self._write_consolidated_csv(results_list, consolidated_csv)
            exported_files.append(consolidated_csv)
            print(f"   ✅ CSV consolidado: {os.path.basename(consolidated_csv)}")
            
            # JSON consolidado
            consolidated_json = os.path.join(exports_dir, "consolidated_data.json")
            self._write_consolidated_json(results_list, consolidated_json)
            exported_files.append(consolidated_json)
            print(f"   ✅ JSON consolidado: {os.path.basename(consolidated_json)}")
            
        except Exception as e:
            print(f"   ❌ Error en exportación consolidada: {e}")
        
        return exported_files
    
    def _write_consolidated_csv(self, results_list: List[AnalysisResults], filename: str):
        """Escribe CSV consolidado."""
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Encabezados
            headers = [
                'Modelo', 'Éxito', 'Tiempo Total (s)',
                'Despl. Máx (m)', 'Convergencia',
                'Primer Período (s)', 'Num. Modos'
            ]
            writer.writerow(headers)
            
            # Datos por modelo
            for result in results_list:
                row = [
                    result.model_name,
                    'Sí' if result.success else 'No',
                    f"{result.total_analysis_time:.3f}"
                ]
                
                # Datos estáticos
                if result.static_results:
                    row.extend([
                        f"{result.static_results.max_displacement:.6f}",
                        'Sí' if result.static_results.convergence_achieved else 'No'
                    ])
                else:
                    row.extend(['N/A', 'N/A'])
                
                # Datos modales
                if result.modal_results and result.modal_results.periods:
                    row.extend([
                        f"{result.modal_results.periods[0]:.4f}",
                        str(len(result.modal_results.periods))
                    ])
                else:
                    row.extend(['N/A', 'N/A'])
                
                writer.writerow(row)
    
    def _write_consolidated_json(self, results_list: List[AnalysisResults], filename: str):
        """Escribe JSON consolidado."""
        consolidated_data = {
            "export_metadata": {
                "exported_at": datetime.now().isoformat(),
                "total_models": len(results_list),
                "export_type": "consolidated",
                "export_version": "1.0.0"
            },
            "summary": {
                "successful_analyses": len([r for r in results_list if r.success]),
                "failed_analyses": len([r for r in results_list if not r.success]),
                "total_analysis_time": sum(r.total_analysis_time for r in results_list)
            },
            "models_data": [
                self._prepare_json_data(result) for result in results_list
            ]
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(consolidated_data, f, indent=2, ensure_ascii=False)
